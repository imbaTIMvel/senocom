import os, sys, re
import pdfplumber
import unicodedata
from pypdf import PdfReader, PdfWriter
from PySide6.QtWidgets import *
from PySide6.QtGui import *
from PySide6.QtCore import *
from openpyxl import load_workbook
from datetime import datetime

# =========================================================
# CONFIGURAÇÕES
# =========================================================

DOCUMENT_TYPES = {
    "Comprovante de Transferência": { "id": 1,
        "anchor": "Comprovante de Transferência",
        "date": ["data da transferencia:"],
        "name": ["nome do recebedor:"],
        "value": ["valor:"],
    },
    "Comprovante de pagamento de boleto": { "id": 2,
        "anchor": "Comprovante de pagamento de boleto",
        "date": ["data de pagamento:"],
        "name": ["razao social:"],
        "value": ["(=) valor do pagamento (r$):"],
    },
    "Comprovante de Transferência CC": { "id": 3,
        "anchor": "conta corrente para conta corrente", # Comprovante de Transferência de conta corrente para conta corrente
        "date": ["transferencia efetuada em"],
        "name": ["nome:"],
        "value": ["valor:"],
    },
    "Comprovante TED": { "id": 4,
        "anchor": "TED solicitada em", # Comprovante de pagamento TED C - outra titularidade
        "date": ["ted solicitada em"],
        "name": ["nome do favorecido:"],
        "value": ["valor da ted:"],
    },
    "Comprovante concessionárias": { "id": 5,
        "anchor": "Comprovante de Pagamento de concessionárias",
        "date": ["operacao efetuada em"],
        "name": ["informacoes fornecidas pelo"],
        "value": ["valor do documento:"],
    },
    "Comprovante DARF": { "id": 6,
        "anchor": "Comprovante de pagamento - DARF",
        "date": ["data do pagamento:"],
        "name": ["agente arrecadador:"],
        "value": ["valor total:"],
    },
    "Comprovante QR Code": { "id": 7,
        "anchor": "Comprovante de pagamento QR Code",
        "date": ["pagamento efetuado em"],
        "name": ["nome do recebedor:"],
        "value": ["valor da transacao:"],
    },
    "Comprovante código de barras": { "id": 8,
        "anchor": "Comprovante de Pagamento com código de barras",
        "date": ["operacao efetuada em"],
        "name": ["informacoes fornecidas pelo"],
        "value": ["valor do documento:"],
    },
    "TED entre bancos": { "id": 9,
        "anchor": "COMPROVANTE DE TRANSFERÊNCIA ENTRE BANCOS", # COMPROVANTE DE TRANSFERÊNCIA ENTRE BANCOS - TED
        "date": ["data da transferencia"],
        "name": ["nome do titular"],
        "value": ["valor da transferencia"],
    },
    "PIX": { "id": 10,
        "anchor": "COMPROVANTE DE PAGAMENTO PIX",
        "date": ["data do pagamento"],
        "name": ["nome do titular"],
        "value": ["valor"],
    },
    "Títulos outros bancos": { "id": 11,
        "anchor": "COMPROVANTE PAGAMENTO TÍTULOS OUTROS BANCOS",
        "date": ["data do pagamento"],
        "name": ["nome do beneficiario"],
        "value": ["valor do pagamento"],
    },
}

INVALID_NAME_TERMS = [
    'cpf',
    'cnpj',
    'conta',
    'valor',
    'r$',
    'pagador'
]

# =========================================================
# UTILITÁRIOS
# =========================================================

def is_valid_name(text):

    text_lower = text.lower()

    if len(text.strip()) < 3:
        return False

    for term in INVALID_NAME_TERMS:
        if term in text_lower:
            return False

    # evita linhas só numéricas
    if re.fullmatch(r'[\d\s\./,-]+', text):
        return False

    return True

def sanitize_filename(text):
    invalid_chars = r'<>:"/\\|?*'

    for char in invalid_chars:
        text = text.replace(char, '')

    text = re.sub(r'\s+', ' ', text).strip()

    return text

def normalize_text(text):
    text = unicodedata.normalize('NFKD', text)
    text = text.encode('ASCII', 'ignore').decode('ASCII')
    text = text.replace('\t', ' ')
    text = re.sub(r' +', ' ', text)
    return text

def extract_value(text):
    match = re.search(r'\d{1,3}(?:\.\d{3})*,\d{2}', text)

    if match:
        return f'R$ {match.group(0)}'

    return 'R$ 0,00'

def extract_date(text):
    match = re.search(r'\d{2}/\d{2}/\d{4}', text)

    if match:
        date = match.group(0)
        parts = date.split('/')
        return f'{parts[0]}-{parts[1]}'

    return 'SEM-DATA'

def normalize_currency(value):

    if value is None:
        return None

    # ==========================================
    # JÁ É NÚMERO (Excel pode devolver float)
    # ==========================================

    if isinstance(value, (int, float)):
        return abs(round(float(value), 2))

    value = str(value).strip()

    # remove R$
    value = value.replace('R$', '').strip()

    # ==========================================
    # FORMATO BRASILEIRO
    # Ex: 22.000,00
    # ==========================================

    if re.fullmatch(r'-?\d{1,3}(?:\.\d{3})*,\d{2}', value):

        value = value.replace('.', '')
        value = value.replace(',', '.')

        try:
            return abs(round(float(value), 2))
        except:
            return None

    # ==========================================
    # FORMATO AMERICANO
    # Ex: 22,000.00
    # ==========================================

    if re.fullmatch(r'-?\d{1,3}(?:,\d{3})*\.\d{2}', value):

        value = value.replace(',', '')

        try:
            return abs(round(float(value), 2))
        except:
            return None

    # ==========================================
    # fallback genérico
    # ==========================================

    try:

        value = value.replace(',', '')

        return abs(round(float(value), 2))

    except:
        return None

def extract_day_month(value):

    # ==========================================
    # DATETIME DO EXCEL
    # ==========================================

    if isinstance(value, datetime):
        return value.strftime('%d-%m')

    text = str(value).strip()

    # ==========================================
    # FORMATO DD/MM/YYYY
    # ==========================================

    match = re.search(
        r'(\d{2})/(\d{2})/\d{4}',
        text
    )

    if match:
        return f'{match.group(1)}-{match.group(2)}'

    # ==========================================
    # FORMATO YYYY-MM-DD
    # ==========================================

    match = re.search(
        r'\d{4}-(\d{2})-(\d{2})',
        text
    )

    if match:
        return f'{match.group(2)}-{match.group(1)}'

    return None

def find_value_after_keyword(lines, keywords):
    for i, line in enumerate(lines):

        line_lower = line.lower()

        for keyword in keywords:

            if keyword.lower() in line_lower:

                # tenta mesma linha
                value = extract_value(line)

                if value != 'R$ 0,00':
                    return value

                # tenta próximas linhas
                for j in range(1, 4):

                    if i + j < len(lines):

                        next_line = lines[i + j]
                        value = extract_value(next_line)

                        if value != 'R$ 0,00':
                            return value

    return 'R$ 0,00'

def find_date_after_keyword(lines, keywords):
    for i, line in enumerate(lines):

        line_lower = line.lower()

        for keyword in keywords:

            if keyword.lower() in line_lower:

                date = extract_date(line)

                if date != 'SEM-DATA':
                    return date

                for j in range(1, 4):

                    if i + j < len(lines):

                        next_line = lines[i + j]
                        date = extract_date(next_line)

                        if date != 'SEM-DATA':
                            return date

    return 'SEM-DATA'

def find_name_after_keyword(lines, keywords, context=None):

    start_index = 0

    # =====================================================
    # CONTEXTO ESPECIAL PIX BRB
    # =====================================================

    if context:

        for i, line in enumerate(lines):

            if context.lower() in line.lower():
                start_index = i
                break

    # =====================================================
    # PROCURA KEYWORDS
    # =====================================================

    for i in range(start_index, len(lines)):

        line = lines[i]
        line_lower = line.lower()

        for keyword in keywords:

            keyword_lower = keyword.lower()

            if keyword_lower in line_lower:

                # =========================================
                # CASO 1 -> INLINE SEM :
                # Ex:
                # Nome do Titular FULANO
                # =========================================

                candidate = re.sub(
                    re.escape(keyword),
                    '',
                    line,
                    flags=re.IGNORECASE
                ).strip()

                candidate = candidate.replace(':', '').strip()

                if (
                    candidate
                    and len(candidate) > 3
                    and is_valid_name(candidate)
                ):
                    return sanitize_filename(candidate)

                # =========================================
                # CASO 2 -> próximas linhas
                # =========================================

                for j in range(1, 4):

                    if i + j < len(lines):

                        next_line = lines[i + j].strip()

                        if (
                            len(next_line) > 2
                            and is_valid_name(next_line)
                        ):
                            return sanitize_filename(next_line)

    return 'SEM-NOME'

def find_line_after_anchor(lines, anchor):
    for i, line in enumerate(lines):

        if anchor.lower() in line.lower():

            if i + 1 < len(lines):

                return sanitize_filename(lines[i + 1].strip())

    return 'SEM-NOME'

def detect_document_type(text):

    text_lower = text.lower()

    matches = []

    for doc_name, config in DOCUMENT_TYPES.items():

        anchor = config['anchor'].lower()

        if anchor in text_lower:
            matches.append((len(anchor), doc_name))
    
    if not matches:
        return None
    
    matches.sort(reverse=True)

    return matches[0][1]

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS

    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# =========================================================
# EXTRAÇÃO PRINCIPAL
# =========================================================

def extract_metadata(text):

    lines = [normalize_text(line.strip()) for line in text.split('\n') if line.strip()]

    doc_type = detect_document_type(text)

    if not doc_type:
        return {
            'id': 0,
            'type': 'DESCONHECIDO',
            'date': 'SEM-DATA',
            'name': 'SEM-NOME',
            'value': 'R$ 0,00'
        }

    config = DOCUMENT_TYPES[doc_type]
    context = None

    if doc_type == 'PIX':
        context = 'Dados de Destino do Pix'
    if doc_type == 'TED entre bancos':
        context = 'Dados de destino'
    if doc_type == 'Comprovante de Transferência CC':
        context = 'Dados da conta creditada'
    if doc_type == 'Comprovante concessionárias':
        name = find_line_after_anchor(lines, normalize_text(config['anchor']))
    else:
        name = find_name_after_keyword(lines, config['name'], context=context)

    doc_id = config['id']
    date = find_date_after_keyword(lines, config['date'])
    value = find_value_after_keyword(lines, config['value'])

    return {
        'id': doc_id,
        'type': doc_type,
        'date': date,
        'name': name,
        'value': value,
    }

# =========================================================
# RELATÓRIO XLSX
# =========================================================

def load_report_data(input_folder):

    xlsx_files = []

    for file in os.listdir(input_folder):

        if file.lower().endswith('.xlsx'):
            xlsx_files.append(
                os.path.join(input_folder, file)
            )

    # ==========================================
    # NENHUM XLSX
    # ==========================================

    if len(xlsx_files) == 0:
        return None

    # ==========================================
    # MAIS DE UM XLSX
    # ==========================================

    if len(xlsx_files) > 1:

        raise Exception(
            'Há mais de um arquivo XLSX na pasta.'
        )

    report_path = xlsx_files[0]

    wb = load_workbook(report_path, data_only=True)

    ws = wb.active

    header_row = None

    # ==========================================
    # PROCURA "Vencimento/Mov."
    # ==========================================

    for row in range(1, 11):

        for col in range(1, ws.max_column + 1):

            value = ws.cell(row=row, column=col).value

            if value and 'Vencimento/Mov.' in str(value):

                header_row = row
                break

        if header_row:
            break

    if not header_row:

        raise Exception(
            'Cabeçalho não encontrado no XLSX.'
        )

    # ==========================================
    # MAPA DE COLUNAS
    # ==========================================

    columns = {}

    for col in range(1, ws.max_column + 1):

        value = ws.cell(
            row=header_row,
            column=col
        ).value

        if value:
            columns[str(value).strip()] = col

    if 'ID' not in columns:

        raise Exception(
            'Coluna "ID" não encontrada.'
        )

    if 'Valor Total' not in columns:

        raise Exception(
            'Coluna "Valor Total" não encontrada.'
        )

    if 'Vencimento/Mov.' not in columns:

        raise Exception(
            'Coluna "Vencimento/Mov." não encontrada.'
        )

    # ==========================================
    # EXTRAÇÃO DAS LINHAS
    # ==========================================

    operations = []

    for row in range(header_row + 1, ws.max_row + 1):

        operation_id = ws.cell(
            row=row,
            column=columns['ID']
        ).value

        operation_date = ws.cell(
            row=row,
            column=columns['Vencimento/Mov.']
        ).value

        operation_value = ws.cell(
            row=row,
            column=columns['Valor Total']
        ).value

        if not operation_id:
            continue

        normalized_date = extract_day_month(
            operation_date
        )

        normalized_value = normalize_currency(
            operation_value
        )

        operations.append({
            'row': row,
            'id': str(operation_id),
            'date': normalized_date,
            'value': normalized_value
        })

    return operations

def find_matching_operation(metadata, operations):

    if not operations:
        return None, []

    pdf_date = metadata['date']

    pdf_value = normalize_currency(
        metadata['value']
    )

    matches = []

    for op in operations:

        if (
            op['date'] == pdf_date
            and op['value'] == pdf_value
        ):
            matches.append(op)

    # ======================================
    # MATCH ÚNICO
    # ======================================

    if len(matches) == 1:
        return matches[0], matches

    return None, matches

# =========================================================
# PDF
# =========================================================

def split_pdf(input_pdf, output_folder, operations=None, error_log=None):

    print(f'Processando: {input_pdf}')

    reader = PdfReader(input_pdf)

    with pdfplumber.open(input_pdf) as pdf:

        page_buffer = []

        for page_number, page in enumerate(pdf.pages):

            try:

                text = page.extract_text()

                if not text:
                    continue

                # guarda página atual no buffer
                page_buffer.append(page_number)

                metadata = extract_metadata(text)

                # ainda não chegou no final do comprovante
                if metadata['type'] == 'DESCONHECIDO':
                    continue

                # =====================================================
                # FINAL DO COMPROVANTE ENCONTRADO
                # =====================================================

                prefix = ''

                # ======================================
                # EXISTE XLSX
                # ======================================

                if operations is not None:

                    matched_operation, matches = find_matching_operation(
                        metadata,
                        operations
                    )

                    # MATCH ÚNICO
                    if matched_operation:

                        raw_id = re.sub(
                            r'\D',
                            '',
                            matched_operation['id']
                        )

                        sid = raw_id.zfill(7)[-7:]

                        prefix = f'SID{sid} '

                    # ERRO
                    else:

                        prefix = '[ERRO] '

                        if error_log is not None:

                            if len(matches) == 0:

                                error_log.append(
                                    (
                                        'SEM MATCH',
                                        metadata,
                                        []
                                    )
                                )

                            else:

                                rows = [m['row'] for m in matches]

                                error_log.append(
                                    (
                                        'MÚLTIPLOS MATCHES',
                                        metadata,
                                        rows
                                    )
                                )

                filename = (
                    # f"[{metadata['id']:02d}] " # RETIRAR DEPOIS - PARA VERSÃO FINAL !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
                    f"{prefix}"
                    f"{metadata['date']} "
                    f"{metadata['name']} "
                    f"{metadata['value']}.pdf"
                )

                filename = sanitize_filename(filename)

                output_path = os.path.join(output_folder, filename)

                # evita sobrescrever
                counter = 1
                original_output_path = output_path

                while os.path.exists(output_path):

                    base, ext = os.path.splitext(original_output_path)

                    output_path = f'{base}_{counter}{ext}'

                    counter += 1

                writer = PdfWriter()

                # salva todas páginas do buffer
                for p in page_buffer:
                    writer.add_page(reader.pages[p])

                with open(output_path, 'wb') as output_file:
                    writer.write(output_file)

                print(f'OK -> {filename}')

                # limpa buffer para próximo comprovante
                page_buffer = []

            except Exception as e:
                print(f'Erro na página {page_number + 1}: {e}')

            # ==============================
            # DEBUGGING
            # ==============================

            # print('\n')
            # print('=' * 80)
            # print(f'PÁGINA {page_number + 1}')
            # print('=' * 80)
            # print(text)
            # print('=' * 80)

# =========================================================
# PROCESSAMENTO EM LOTE
# =========================================================

def process_folder(input_folder, output_folder):

    pdf_files = []

    for root, dirs, files in os.walk(input_folder):

        for file in files:

            if file.lower().endswith('.pdf'):
                pdf_files.append(os.path.join(root, file))

    if not pdf_files:
        QMessageBox.warning(window, 'Aviso', 'Nenhum PDF encontrado.')
        return
    
    # ==================================================
    # CARREGA RELATÓRIO
    # ==================================================

    operations = load_report_data(input_folder)
    error_log = []
    for pdf_file in pdf_files:
        split_pdf(pdf_file, output_folder, operations, error_log)
    
    # ==================================================
    # GERA TXT DE ERROS
    # ==================================================

    if error_log:
        log_path = os.path.join(
            output_folder,
            'erros_do_algoritmo.txt'
        )
        with open(
            log_path,
            'w',
            encoding='utf-8'
        ) as f:
            for error_type, metadata, rows in error_log:
                f.write(
                    f'{error_type}\n'
                )
                f.write(
                    f'Arquivo: '
                    f'{metadata["date"]} '
                    f'{metadata["name"]} '
                    f'{metadata["value"]}\n'
                )
                if rows:
                    f.write(
                        f'Linhas possíveis: {rows}\n'
                    )
                f.write('\n')

# =========================================================
# THREAD DE PROCESSAMENTO
# =========================================================

class Worker(QThread):

    progress = Signal(int)
    finish_proc = Signal()

    def __init__(self, input_folder, output_folder):
        super().__init__()

        self.input_folder = input_folder
        self.output_folder = output_folder

    def run(self):

        try:

            process_folder(
                self.input_folder,
                self.output_folder
            )

        except Exception as e:

            QMessageBox.critical(
                window,
                'Erro',
                str(e)
            )

        self.progress.emit(100)

        self.finish_proc.emit()

# =========================================================
# WINDOW
# =========================================================

app = QApplication(sys.argv)

window = QWidget()
window.setWindowTitle("SeNoCom")
window.resize(700, 400)
window.setWindowIcon(QIcon(resource_path("senocom.ico")))

# =========================================================
# BACKGROUND
# =========================================================

window.setStyleSheet("""
QWidget {
    background-color: #1e1e1e;
}
""")

# =========================================================
# BACKGROUND IMAGE
# =========================================================

bg_label = QLabel(window)
bg_pixmap = QPixmap(resource_path("bg_hbr.png"))
bg_label.setPixmap(bg_pixmap)
bg_label.setScaledContents(True)

# =========================================================
# CARD
# =========================================================

card = QFrame()
card.setObjectName("card")

card.setStyleSheet("""
#card {
    background-color: rgba(30,30,30,220);
    border-radius: 16px;
}
""")

# =========================================================
# TITLE
# =========================================================

titulo = QLabel("SeNoCom")

titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)

titulo.setStyleSheet("""
font-family: "Bahnschrift Condensed";
font-size: 38px;
font-weight: bold;
color: white;
padding-bottom: 10px;
""")

# =========================================================
# INPUTS
# =========================================================

input_style = """
QLineEdit {
    background-color: #2b2b2b;
    color: white;
    border: 1px solid #3a3a3a;
    border-radius: 8px;
    padding: 8px;
    font-size: 14px;
}
"""

entry_input = QLineEdit()
entry_input.setPlaceholderText("Selecione a pasta contendo os PDFs...")
entry_input.setStyleSheet(input_style)

# =========================================================
# BUTTON STYLE
# =========================================================

button_style = """
QPushButton {
    background-color: #f9b02e;
    color: black;
    border: none;
    border-radius: 8px;
    padding: 10px;
    font-weight: bold;
    font-size: 14px;
}

QPushButton:hover {
    background-color: #ffd166;
}

QPushButton:pressed {
    background-color: #e69500;
}

QPushButton:disabled {
    background-color: #666666;
    color: #aaaaaa;
}
"""

# =========================================================
# FILE DIALOGS
# =========================================================

def selecionar_entrada():

    folder = QFileDialog.getExistingDirectory(
        window,
        "Selecione a pasta de entrada"
    )

    if folder:
        entry_input.setText(folder)

# =========================================================
# BUTTONS
# =========================================================

btn_input = QPushButton("Selecionar Entrada")
btn_input.setStyleSheet(button_style)
btn_input.clicked.connect(selecionar_entrada)

btn_start = QPushButton("Executar Processamento")
btn_start.setStyleSheet(button_style)

# =========================================================
# PROGRESS BAR
# =========================================================

progress = QProgressBar()

progress.setStyleSheet("""
QProgressBar {
    background-color: #2b2b2b;
    border-radius: 6px;
    text-align: center;
    color: white;
    height: 18px;
}

QProgressBar::chunk {
    background-color: #f9b02e;
    border-radius: 6px;
}
""")

# =========================================================
# STATUS
# =========================================================

label_status = QLabel("")

label_status.setAlignment(Qt.AlignmentFlag.AlignCenter)

label_status.setStyleSheet("""
color: #cccccc;
font-size: 12px;
padding-top: 5px;
""")

# =========================================================
# EXECUÇÃO
# =========================================================

worker = None

def iniciar_processamento():

    global worker

    input_folder = entry_input.text().strip()

    if not input_folder:
        QMessageBox.warning(
            window,
            'Aviso',
            'Selecione a pasta de entrada.'
        )
        return

    # ============================================
    # ESCOLHE SAÍDA SOMENTE AO FINAL
    # ============================================

    output_folder = QFileDialog.getExistingDirectory(
        window,
        "Selecione a pasta de destino"
    )

    if not output_folder:
        return

    progress.setValue(0)

    btn_start.setEnabled(False)

    label_status.setText('Processando arquivos...')

    worker = Worker(input_folder, output_folder)

    worker.progress.connect(progress.setValue)

    worker.finish_proc.connect(processamento_finalizado)

    worker.start()

def processamento_finalizado():

    print("PROCESSAMENTO FINALIZADO") ##################################################################

    btn_start.setEnabled(True)

    label_status.setText('Processamento concluído.')

    QMessageBox.information(
        window,
        'Concluído',
        'Arquivos processados com sucesso.'
    )

btn_start.clicked.connect(iniciar_processamento)

# ==============================
# FOOTER / ASSINATURA
# ==============================

github_label = QLabel()
github_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

github_label.setText(
    '<a href="https://github.com/imbaTIMvel/senocom">'
    'SeNoCom v0.1.0 - GitHub'
    '</a>'
)

github_label.setOpenExternalLinks(True)

github_label.setStyleSheet("""
QLabel {
    background-color: transparent;
    color: rgba(255,255,255,120);
    font-size: 11px;
}

QLabel:hover {
    color: #f9b02e;
}
""")

footer = QLabel(
    "Desenvolvido por: Diretoria Administrativa Financeira - DAF"
)

footer.setAlignment(Qt.AlignmentFlag.AlignCenter)

footer.setStyleSheet("""
QLabel {
    background-color: transparent;
    color: rgba(255,255,255,120);
    font-size: 10px;
    padding-bottom: 4px;
}
""")

# =========================================================
# LAYOUTS
# =========================================================

layout_input = QHBoxLayout()
layout_input.addWidget(entry_input)
layout_input.addWidget(btn_input)

card_layout = QVBoxLayout(card)

card_layout.addWidget(titulo)
card_layout.addLayout(layout_input)
card_layout.addWidget(btn_start)
card_layout.addWidget(progress)
card_layout.addWidget(label_status)

card_layout.setSpacing(15)
card_layout.setContentsMargins(25, 25, 25, 25)
card.setMaximumWidth(650)

main_layout = QVBoxLayout(window)

main_layout.addStretch()
main_layout.addWidget(card, alignment=Qt.AlignmentFlag.AlignCenter
)
main_layout.addStretch()
main_layout.addWidget(github_label, alignment=Qt.AlignmentFlag.AlignBottom)
main_layout.addWidget(footer, alignment=Qt.AlignmentFlag.AlignBottom)

main_layout.setContentsMargins(40, 40, 40, 40)

window.setLayout(main_layout)

# =========================================================
# RESPONSIVE BACKGROUND
# =========================================================

def resize_event(event):

    bg_label.resize(window.size())

window.resizeEvent = resize_event

# =========================================================
# SHOW
# =========================================================

window.show()

sys.exit(app.exec())
